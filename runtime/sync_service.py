"""공유 워크스페이스 기준 sync/review/export를 조율한다."""

from __future__ import annotations

from dataclasses import asdict, dataclass, field
from datetime import datetime
import json
import os
from pathlib import Path
import shutil
from typing import Any

from openpyxl import Workbook, load_workbook

from analysis import run_inbox_review_board_smoke
from exports import (
    ProjectedTemplateRow,
    append_projected_row_to_workbook,
    apply_hybrid_template_mapping,
    read_template_profile,
)
from llm import OpenAIResponsesConfig, OpenAIResponsesWrapper
from mailbox import build_local_mailbox_account_config, run_imap_inbox_backfill_smoke
from mailbox.imap_backfill_smoke import default_backfill_report_path
from runtime.lockfile import WorkspaceWriteLockHandle, acquire_workspace_write_lock
from runtime.review_state import ingest_review_report_into_state
from runtime.secrets_store import WorkspaceSecretsStore
from runtime.state_store import WorkspaceStateStore
from runtime.workspace import SharedWorkspace, load_shared_workspace


@dataclass(slots=True)
class WorkspaceSyncResult:
    """기능: 공유 워크스페이스 sync 1회의 결과를 표현한다."""

    workspace_root: str
    profile_root: str
    backfill_report_path: str
    review_json_path: str
    review_html_path: str
    operating_workbook_path: str
    representative_export_count: int
    total_review_item_count: int
    sync_run_id: int
    notes: list[str] = field(default_factory=list)

    def to_dict(self) -> dict[str, object]:
        return asdict(self)


def run_workspace_sync(
    *,
    workspace_root: str | Path,
    workspace_password: str,
    app_kind: str = "server-tool",
    profile_id: str = "shared-workspace",
    force_lock_takeover: bool = False,
) -> WorkspaceSyncResult:
    """기능: 공유 워크스페이스에서 backfill -> review -> 운영 workbook 재구성을 수행한다."""

    workspace = load_shared_workspace(workspace_root)
    state_store = WorkspaceStateStore(workspace.state_db_path())
    state_store.ensure_schema()
    secrets_store = WorkspaceSecretsStore(
        path=str(workspace.secure_blob_path()),
        password=workspace_password,
    )
    secrets_payload = secrets_store.read()

    write_lock = acquire_workspace_write_lock(
        lock_path=workspace.lock_path(),
        workspace_id=workspace.manifest.workspace_id,
        app_kind=app_kind,
        force_takeover=force_lock_takeover,
    )
    sync_run_id = state_store.start_sync_run(
        run_kind="workspace_sync",
        app_kind=app_kind,
        metadata={"workspace_id": workspace.manifest.workspace_id},
    )
    try:
        mailbox_settings = dict(secrets_payload.get("mailbox") or {})
        llm_settings = dict(secrets_payload.get("llm") or {})
        export_settings = dict(secrets_payload.get("exports") or {})

        account_config = build_local_mailbox_account_config(
            email_address=str(mailbox_settings.get("email_address") or ""),
            login_username=str(mailbox_settings.get("login_username") or ""),
            password=str(mailbox_settings.get("password") or ""),
            profile_root=workspace.profile_root(),
            source_path="workspace.encrypted_settings",
            notes=["공유 워크스페이스 암호화 설정에서 메일 계정 정보를 읽었다."],
        )
        template_path = _resolve_template_workbook_path(
            workspace=workspace,
            export_settings=export_settings,
        )
        wrapper = OpenAIResponsesWrapper(
            OpenAIResponsesConfig(
                model=str(llm_settings.get("model") or "gpt-5.4"),
                api_key=str(llm_settings.get("api_key") or ""),
                usage_log_path=str(workspace.profile_paths().llm_usage_log_path()),
            )
        )

        backfill_report = run_imap_inbox_backfill_smoke(
            account_config=account_config,
            folder=str(mailbox_settings.get("default_folder") or "INBOX"),
        )
        write_lock.refresh()

        review_report = run_inbox_review_board_smoke(
            profile_id=profile_id,
            profile_root=str(workspace.profile_root()),
            template_path=str(template_path),
            reuse_existing_analysis=False,
            wrapper=wrapper,
        )
        _update_latest_review_pointers(workspace=workspace, review_report=review_report)
        write_lock.refresh()

        review_items = ingest_review_report_into_state(
            workspace=workspace,
            state_store=state_store,
            report_path=review_report.review_json_path,
        )
        workbook_result = rebuild_operating_workbook(
            workspace=workspace,
            state_store=state_store,
            template_path=template_path,
            wrapper=wrapper,
        )
        write_lock.refresh()

        notes = [
            "공유 워크스페이스 기준 backfill, triage review, 운영 workbook 재구성을 완료했다.",
            f"backfill_success={backfill_report.success}",
        ]
        result = WorkspaceSyncResult(
            workspace_root=str(workspace.root()),
            profile_root=str(workspace.profile_root()),
            backfill_report_path=str(
                workspace.to_workspace_relative(
                    default_backfill_report_path(
                        str(workspace.profile_root()),
                        account_config.email_address,
                    )
                )
            ),
            review_json_path=str(workspace.to_workspace_relative(review_report.review_json_path)),
            review_html_path=str(workspace.to_workspace_relative(review_report.review_html_path)),
            operating_workbook_path=str(workbook_result["operating_workbook_relpath"]),
            representative_export_count=int(workbook_result["representative_count"]),
            total_review_item_count=len(review_items),
            sync_run_id=sync_run_id,
            notes=notes,
        )
        state_store.finish_sync_run(
            sync_run_id,
            status="completed",
            notes=result.notes,
            metadata=result.to_dict(),
        )
        return result
    except Exception as exc:
        state_store.finish_sync_run(
            sync_run_id,
            status="failed",
            notes=[f"{exc.__class__.__name__}: {exc}"],
            metadata={"workspace_root": str(workspace.root())},
        )
        raise
    finally:
        write_lock.release()


def rebuild_operating_workbook(
    *,
    workspace: SharedWorkspace,
    state_store: WorkspaceStateStore,
    template_path: str | Path,
    wrapper: OpenAIResponsesWrapper,
) -> dict[str, Any]:
    """기능: state DB 기준 대표 신청 건만 stable 운영 workbook으로 다시 쓴다."""

    profile_paths = workspace.profile_paths()
    operating_path = profile_paths.operating_export_workbook_path()
    snapshot_root = profile_paths.runtime_exports_snapshots_root()
    snapshot_root.mkdir(parents=True, exist_ok=True)
    if operating_path.exists():
        snapshot_name = datetime.now().strftime("%y%m%d_%H%M") + "_기업_신청서_모음.xlsx"
        shutil.copy2(operating_path, snapshot_root / snapshot_name)

    template_profile = read_template_profile(
        workbook_path=str(template_path),
        profile_id="shared-workspace",
        template_id=Path(template_path).stem,
    )
    mapped_profile, _ = apply_hybrid_template_mapping(template_profile, wrapper=wrapper)
    _prepare_empty_output_workbook(
        template_path=Path(template_path),
        output_workbook_path=operating_path,
        template_profile=mapped_profile,
    )

    representative_items = state_store.list_review_items(export_only=True)
    workbook_row_items: list[dict[str, Any]] = []
    for item in representative_items:
        projected_row_relpath = item.get("projected_row_relpath")
        if not projected_row_relpath:
            continue
        projected_row = ProjectedTemplateRow.from_dict(
            json.loads(
                workspace.from_workspace_relative(projected_row_relpath).read_text(
                    encoding="utf-8"
                )
            )
        )
        append_result = append_projected_row_to_workbook(
            profile=mapped_profile,
            projected_row=projected_row,
            output_workbook_path=str(operating_path),
            source_workbook_path=str(template_path),
        )
        workbook_row_items.append(
            {
                "bundle_id": item["bundle_id"],
                "sheet_name": append_result.sheet_name,
                "row_index": append_result.appended_row_index,
            }
        )

    workbook_relpath = workspace.to_workspace_relative(operating_path)
    if workbook_row_items:
        state_store.save_workbook_rows(
            workbook_relpath=workbook_relpath,
            row_items=workbook_row_items,
        )
    else:
        state_store.clear_workbook_rows()

    refreshed_items = state_store.list_review_items()
    _write_review_index_sheet(
        workspace=workspace,
        workbook_path=operating_path,
        items=refreshed_items,
    )
    return {
        "operating_workbook_relpath": workbook_relpath,
        "representative_count": len(workbook_row_items),
    }


def _resolve_template_workbook_path(
    *,
    workspace: SharedWorkspace,
    export_settings: dict[str, Any],
) -> Path:
    relative_path = str(export_settings.get("template_workbook_relative_path") or "").strip()
    if relative_path:
        return workspace.from_workspace_relative(relative_path)
    default_template = workspace.profile_paths().template_workbook_path()
    if default_template.exists():
        return default_template
    raise RuntimeError("공유 워크스페이스에 template workbook 경로가 없다.")


def _prepare_empty_output_workbook(
    *,
    template_path: Path,
    output_workbook_path: Path,
    template_profile,
) -> None:
    workbook = load_workbook(template_path)
    for sheet in template_profile.sheets:
        worksheet = workbook[sheet.sheet_name]
        for row_index in range(sheet.data_start_row_index, worksheet.max_row + 1):
            for column_index in range(1, worksheet.max_column + 1):
                cell = worksheet.cell(row=row_index, column=column_index)
                cell.value = None
                cell.hyperlink = None
    output_workbook_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_workbook_path)


def _write_review_index_sheet(
    *,
    workspace: SharedWorkspace,
    workbook_path: Path,
    items: list[dict[str, Any]],
) -> None:
    workbook = load_workbook(workbook_path)
    if "검토_인덱스" in workbook.sheetnames:
        worksheet = workbook["검토_인덱스"]
        workbook.remove(worksheet)
    worksheet = workbook.create_sheet("검토_인덱스")
    headers = [
        "received_at",
        "sender",
        "subject",
        "triage",
        "export_status",
        "dedupe_status",
        "company_name",
        "bundle_id",
        "workbook_row",
        "preview_path",
        "raw_eml_path",
    ]
    for column_index, header in enumerate(headers, start=1):
        worksheet.cell(row=1, column=column_index).value = header

    workbook_dir = workbook_path.parent
    for row_index, item in enumerate(items, start=2):
        dedupe_status = "representative" if item["is_export_representative"] else (
            "duplicate_application" if item["duplicate_of_bundle_id"] else ""
        )
        preview_rel = item.get("preview_relpath") or ""
        raw_eml_rel = item.get("raw_eml_relpath") or ""
        values = [
            item.get("received_at") or "",
            item.get("sender") or "",
            item.get("subject") or "",
            item.get("triage_label") or "",
            item.get("export_status") or "",
            dedupe_status,
            item.get("company_name") or "",
            item.get("bundle_id") or "",
            item.get("workbook_row_index") or "",
            preview_rel,
            raw_eml_rel,
        ]
        for column_index, value in enumerate(values, start=1):
            cell = worksheet.cell(row=row_index, column=column_index)
            cell.value = value
            if column_index in {10, 11} and value:
                target = workspace.from_workspace_relative(str(value))
                try:
                    relative_target = os.path.relpath(target, workbook_dir)
                    cell.hyperlink = relative_target
                except Exception:
                    pass
    workbook.save(workbook_path)


def _update_latest_review_pointers(*, workspace: SharedWorkspace, review_report) -> None:
    review_root = workspace.review_logs_root()
    review_root.mkdir(parents=True, exist_ok=True)
    latest_json = review_root / "latest_inbox_review_board.json"
    latest_html = review_root / "latest_inbox_review_board.html"
    shutil.copy2(review_report.review_json_path, latest_json)
    shutil.copy2(review_report.review_html_path, latest_html)
