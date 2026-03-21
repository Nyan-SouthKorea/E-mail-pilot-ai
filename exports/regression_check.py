"""생성 workbook과 reference fixture workbook의 차이를 비교한다."""

from __future__ import annotations

import argparse
import json
from dataclasses import asdict, dataclass, field
from pathlib import Path

from openpyxl import load_workbook
from project_paths import ProfilePaths, default_example_profile_root

from .rule_mapping import apply_rule_based_template_mapping
from .schema import TemplateColumn, TemplateProfile, TemplateSheet
from .template_profile import read_template_profile


@dataclass(slots=True)
class WorkbookCellDiff:
    """기능: workbook 두 셀 사이의 차이를 표현한다.

    입력:
    - column_index: 1-based 열 번호
    - column_letter: Excel 열 문자
    - header_text: 헤더 이름
    - expected_value: reference workbook 값
    - actual_value: generated workbook 값

    반환:
    - dataclass 인스턴스
    """

    column_index: int
    column_letter: str
    header_text: str
    expected_value: str
    actual_value: str

    def to_dict(self) -> dict[str, object]:
        """기능: 차이 정보를 JSON 직렬화용 dict로 바꾼다.

        입력:
        - 없음

        반환:
        - dict
        """

        return asdict(self)


@dataclass(slots=True)
class WorkbookRowComparison:
    """기능: reference row와 generated row 한 쌍의 비교 결과를 표현한다.

    입력:
    - reference_row_index: reference workbook 행 번호
    - generated_row_index: generated workbook 행 번호
    - compared_cell_count: 비교한 셀 개수
    - matched_cell_count: 일치한 셀 개수
    - differing_cells: 차이가 난 셀 목록

    반환:
    - dataclass 인스턴스
    """

    reference_row_index: int
    generated_row_index: int
    compared_cell_count: int
    matched_cell_count: int
    differing_cells: list[WorkbookCellDiff] = field(default_factory=list)

    def match_ratio(self) -> float:
        """기능: 현재 행의 셀 일치 비율을 반환한다.

        입력:
        - 없음

        반환:
        - 0.0~1.0 비율
        """

        if self.compared_cell_count == 0:
            return 1.0
        return self.matched_cell_count / self.compared_cell_count

    def to_dict(self) -> dict[str, object]:
        """기능: 비교 결과를 JSON 직렬화용 dict로 바꾼다.

        입력:
        - 없음

        반환:
        - dict
        """

        payload = asdict(self)
        payload["match_ratio"] = round(self.match_ratio(), 4)
        return payload


@dataclass(slots=True)
class WorkbookRegressionReport:
    """기능: workbook 회귀 비교 결과 전체를 표현한다.

    입력:
    - reference_workbook_path: 기준 workbook 경로
    - generated_workbook_path: 생성 workbook 경로
    - sheet_name: 비교 대상 시트명
    - compared_row_count: 비교한 행 수
    - compared_cell_count: 비교한 셀 수
    - matched_cell_count: 일치한 셀 수
    - ignored_headers: 비교에서 제외한 헤더 목록
    - row_comparisons: 행 단위 비교 목록
    - notes: 추가 메모

    반환:
    - dataclass 인스턴스
    """

    reference_workbook_path: str
    generated_workbook_path: str
    sheet_name: str
    compared_row_count: int
    compared_cell_count: int
    matched_cell_count: int
    ignored_headers: list[str] = field(default_factory=list)
    row_comparisons: list[WorkbookRowComparison] = field(default_factory=list)
    notes: list[str] = field(default_factory=list)

    def match_ratio(self) -> float:
        """기능: 전체 셀 일치 비율을 반환한다.

        입력:
        - 없음

        반환:
        - 0.0~1.0 비율
        """

        if self.compared_cell_count == 0:
            return 1.0
        return self.matched_cell_count / self.compared_cell_count

    def differing_cell_count(self) -> int:
        """기능: 전체 차이 셀 수를 반환한다.

        입력:
        - 없음

        반환:
        - 차이 셀 개수
        """

        return self.compared_cell_count - self.matched_cell_count

    def to_dict(self) -> dict[str, object]:
        """기능: 회귀 보고서를 JSON 직렬화용 dict로 바꾼다.

        입력:
        - 없음

        반환:
        - dict
        """

        payload = asdict(self)
        payload["match_ratio"] = round(self.match_ratio(), 4)
        payload["differing_cell_count"] = self.differing_cell_count()
        return payload


def compare_fixture_workbooks(
    *,
    reference_workbook_path: str,
    generated_workbook_path: str,
    sheet_name: str | None = None,
    ignored_headers: list[str] | None = None,
) -> WorkbookRegressionReport:
    """기능: reference workbook과 generated workbook의 fixture 결과를 비교한다.

    입력:
    - reference_workbook_path: 기대 산출물 workbook 경로
    - generated_workbook_path: 생성된 workbook 경로
    - sheet_name: 특정 시트만 비교할 때의 시트명
    - ignored_headers: 비교에서 제외할 헤더 목록

    반환:
    - `WorkbookRegressionReport`
    """

    ignored_headers = ignored_headers or ["번호"]

    template_profile = read_template_profile(
        workbook_path=reference_workbook_path,
        profile_id="regression-check",
        template_id=Path(reference_workbook_path).stem,
    )
    mapped_profile, _ = apply_rule_based_template_mapping(template_profile)
    template_sheet = _select_template_sheet(mapped_profile, sheet_name=sheet_name)

    reference_workbook = load_workbook(reference_workbook_path, data_only=True)
    generated_workbook = load_workbook(generated_workbook_path, data_only=True)
    reference_sheet = reference_workbook[template_sheet.sheet_name]
    generated_sheet = generated_workbook[template_sheet.sheet_name]

    comparable_columns = [
        column
        for column in template_sheet.columns
        if column.header_text not in ignored_headers
    ]
    reference_rows = _collect_data_rows(reference_sheet, template_sheet, comparable_columns)
    generated_rows = _collect_data_rows(generated_sheet, template_sheet, comparable_columns)

    notes: list[str] = []
    if len(generated_rows) < len(reference_rows):
        notes.append("generated workbook의 비교 대상 실데이터 행 수가 reference보다 적다.")

    compared_row_count = min(len(reference_rows), len(generated_rows))
    aligned_generated_rows = generated_rows[-compared_row_count:] if compared_row_count else []
    aligned_reference_rows = reference_rows[:compared_row_count]

    row_comparisons: list[WorkbookRowComparison] = []
    compared_cell_count = 0
    matched_cell_count = 0

    for reference_row_index, generated_row_index in zip(
        aligned_reference_rows,
        aligned_generated_rows,
    ):
        row_comparison = _compare_row_pair(
            reference_sheet=reference_sheet,
            generated_sheet=generated_sheet,
            reference_row_index=reference_row_index,
            generated_row_index=generated_row_index,
            columns=comparable_columns,
        )
        row_comparisons.append(row_comparison)
        compared_cell_count += row_comparison.compared_cell_count
        matched_cell_count += row_comparison.matched_cell_count

    if len(generated_rows) > len(reference_rows):
        notes.append(
            "generated workbook의 마지막 실데이터 행들을 reference fixture 행과 대응시켜 비교했다."
        )

    return WorkbookRegressionReport(
        reference_workbook_path=reference_workbook_path,
        generated_workbook_path=generated_workbook_path,
        sheet_name=template_sheet.sheet_name,
        compared_row_count=compared_row_count,
        compared_cell_count=compared_cell_count,
        matched_cell_count=matched_cell_count,
        ignored_headers=list(ignored_headers),
        row_comparisons=row_comparisons,
        notes=notes,
    )


def save_regression_report(
    report: WorkbookRegressionReport,
    output_path: str,
) -> Path:
    """기능: 회귀 비교 보고서를 JSON 파일로 저장한다.

    입력:
    - report: 회귀 비교 결과
    - output_path: 저장 경로

    반환:
    - 저장된 `Path`
    """

    path = Path(output_path)
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(
        json.dumps(report.to_dict(), ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    return path


def default_reference_workbook_path() -> Path:
    """기능: 현재 fixture 기준 reference workbook 경로를 반환한다.

    입력:
    - 없음

    반환:
    - reference workbook 경로
    """

    profile_paths = ProfilePaths(str(default_example_profile_root()))
    return profile_paths.template_workbook_path()


def default_generated_workbook_path() -> Path:
    """기능: 현재 fixture 기준 generated workbook 경로를 반환한다.

    입력:
    - 없음

    반환:
    - generated workbook 경로
    """

    profile_paths = ProfilePaths(str(default_example_profile_root()))
    return profile_paths.runtime_exports_root() / "기업 신청서 모음_fixture_pipeline.xlsx"


def default_regression_report_path() -> Path:
    """기능: 현재 fixture 기준 regression report 저장 경로를 반환한다.

    입력:
    - 없음

    반환:
    - JSON report 경로
    """

    profile_paths = ProfilePaths(str(default_example_profile_root()))
    return profile_paths.runtime_exports_logs_root() / "fixture_regression_report.json"


def main() -> None:
    """기능: CLI에서 workbook 회귀 비교를 실행한다.

    입력:
    - 없음

    반환:
    - 없음
    """

    parser = argparse.ArgumentParser(description="fixture workbook regression check")
    parser.add_argument(
        "--reference-workbook-path",
        default=str(default_reference_workbook_path()),
    )
    parser.add_argument(
        "--generated-workbook-path",
        default=str(default_generated_workbook_path()),
    )
    parser.add_argument(
        "--output-report-path",
        default=str(default_regression_report_path()),
    )
    parser.add_argument(
        "--sheet-name",
        default="",
    )
    parser.add_argument(
        "--ignore-header",
        action="append",
        default=[],
        help="비교에서 제외할 헤더를 추가한다. 여러 번 사용할 수 있다.",
    )
    args = parser.parse_args()

    report = compare_fixture_workbooks(
        reference_workbook_path=args.reference_workbook_path,
        generated_workbook_path=args.generated_workbook_path,
        sheet_name=args.sheet_name or None,
        ignored_headers=args.ignore_header or None,
    )
    save_regression_report(report, args.output_report_path)
    print(json.dumps(report.to_dict(), ensure_ascii=False, indent=2))


def _select_template_sheet(
    profile: TemplateProfile,
    sheet_name: str | None,
) -> TemplateSheet:
    if sheet_name:
        for sheet in profile.sheets:
            if sheet.sheet_name == sheet_name:
                return sheet
        raise RuntimeError(f"시트 `{sheet_name}`를 찾지 못했다.")

    sheet = profile.primary_sheet()
    if sheet is None:
        raise RuntimeError("비교 대상 템플릿 시트를 찾지 못했다.")
    return sheet


def _collect_data_rows(
    worksheet,
    template_sheet: TemplateSheet,
    columns: list[TemplateColumn],
) -> list[int]:
    row_indexes: list[int] = []
    column_indexes = [column.column_index for column in columns]
    for row_index in range(template_sheet.data_start_row_index, worksheet.max_row + 1):
        if _row_has_any_value(worksheet, row_index, column_indexes):
            row_indexes.append(row_index)
    return row_indexes


def _row_has_any_value(worksheet, row_index: int, column_indexes: list[int]) -> bool:
    for column_index in column_indexes:
        value = worksheet.cell(row=row_index, column=column_index).value
        if _normalize_cell_value(value):
            return True
    return False


def _compare_row_pair(
    *,
    reference_sheet,
    generated_sheet,
    reference_row_index: int,
    generated_row_index: int,
    columns: list[TemplateColumn],
) -> WorkbookRowComparison:
    differing_cells: list[WorkbookCellDiff] = []
    compared_cell_count = 0
    matched_cell_count = 0

    for column in columns:
        compared_cell_count += 1
        expected_value = _normalize_cell_value(
            reference_sheet.cell(row=reference_row_index, column=column.column_index).value
        )
        actual_value = _normalize_cell_value(
            generated_sheet.cell(row=generated_row_index, column=column.column_index).value
        )
        if expected_value == actual_value:
            matched_cell_count += 1
            continue

        differing_cells.append(
            WorkbookCellDiff(
                column_index=column.column_index,
                column_letter=column.column_letter,
                header_text=column.header_text,
                expected_value=expected_value,
                actual_value=actual_value,
            )
        )

    return WorkbookRowComparison(
        reference_row_index=reference_row_index,
        generated_row_index=generated_row_index,
        compared_cell_count=compared_cell_count,
        matched_cell_count=matched_cell_count,
        differing_cells=differing_cells,
    )


def _normalize_cell_value(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, str):
        return value.strip()
    return str(value).strip()


if __name__ == "__main__":
    main()
