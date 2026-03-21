"""투영된 템플릿 행을 실제 workbook에 append 한다."""

from __future__ import annotations

from copy import copy
from dataclasses import dataclass, field
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.formula.translate import Translator

from .record_projection import ProjectedTemplateRow, ProjectedTemplateValue
from .schema import TemplateProfile, TemplateSheet


@dataclass(slots=True)
class WorkbookAppendResult:
    """기능: workbook append 실행 결과를 표현한다.

    입력:
    - output_workbook_path: 실제 저장된 workbook 경로
    - sheet_name: 값을 쓴 시트명
    - appended_row_index: 값을 쓴 행 번호
    - last_data_row_index: append 직전 마지막 실데이터 행
    - inherited_style_row_index: 서식/수식을 복사한 기준 행
    - written_cells: 실제 값이 들어간 셀 좌표 목록
    - notes: 추가 메모

    반환:
    - dataclass 인스턴스
    """

    output_workbook_path: str
    sheet_name: str
    appended_row_index: int
    last_data_row_index: int
    inherited_style_row_index: int | None = None
    written_cells: list[str] = field(default_factory=list)
    notes: list[str] = field(default_factory=list)


def append_projected_row_to_workbook(
    profile: TemplateProfile,
    projected_row: ProjectedTemplateRow,
    output_workbook_path: str,
    source_workbook_path: str | None = None,
) -> WorkbookAppendResult:
    """기능: 투영된 템플릿 행을 workbook에 append하고 결과 파일을 저장한다.

    입력:
    - profile: 템플릿 구조와 원본 workbook 경로를 가진 프로필
    - projected_row: 템플릿 열 기준으로 값이 준비된 한 줄
    - output_workbook_path: 생성 또는 갱신할 결과 workbook 경로
    - source_workbook_path: 출력본이 아직 없을 때 읽을 원본 workbook 경로

    반환:
    - `WorkbookAppendResult`
    """

    output_path = Path(output_workbook_path)
    source_path = Path(source_workbook_path or profile.source_workbook_path)
    workbook_path = output_path if output_path.exists() else source_path

    workbook = load_workbook(workbook_path)
    template_sheet = _select_template_sheet(profile, projected_row.sheet_name)
    worksheet = workbook[template_sheet.sheet_name]

    last_data_row_index = _find_last_data_row(worksheet, template_sheet)
    appended_row_index = max(last_data_row_index + 1, template_sheet.data_start_row_index)
    inherited_style_row_index = _select_style_source_row(
        template_sheet=template_sheet,
        last_data_row_index=last_data_row_index,
    )

    if inherited_style_row_index is not None and inherited_style_row_index != appended_row_index:
        _copy_row_presentation(
            worksheet=worksheet,
            source_row_index=inherited_style_row_index,
            target_row_index=appended_row_index,
        )

    written_cells: list[str] = []
    for projected_value in projected_row.values:
        cell = worksheet.cell(
            row=appended_row_index,
            column=projected_value.column_index,
        )
        cell.value = _resolve_output_value(
            projected_value=projected_value,
            worksheet=worksheet,
            template_sheet=template_sheet,
            append_row_index=appended_row_index,
            last_data_row_index=last_data_row_index,
        )
        written_cells.append(cell.coordinate)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)

    notes: list[str] = []
    if output_path.exists() and workbook_path == output_path:
        notes.append("기존 결과 workbook에 새 행을 이어서 append했다.")
    else:
        notes.append("원본 템플릿을 읽어 결과 workbook을 새로 생성했다.")

    return WorkbookAppendResult(
        output_workbook_path=str(output_path),
        sheet_name=template_sheet.sheet_name,
        appended_row_index=appended_row_index,
        last_data_row_index=last_data_row_index,
        inherited_style_row_index=inherited_style_row_index,
        written_cells=written_cells,
        notes=notes,
    )


def _select_template_sheet(
    profile: TemplateProfile,
    sheet_name: str,
) -> TemplateSheet:
    for sheet in profile.sheets:
        if sheet.sheet_name == sheet_name:
            return sheet

    primary_sheet = profile.primary_sheet()
    if primary_sheet is not None:
        return primary_sheet

    raise RuntimeError("append 대상 템플릿 시트를 찾지 못했다.")


def _find_last_data_row(worksheet, template_sheet: TemplateSheet) -> int:
    last_data_row_index = template_sheet.data_start_row_index - 1
    relevant_columns = [column.column_index for column in template_sheet.columns]

    for row_index in range(template_sheet.data_start_row_index, worksheet.max_row + 1):
        if _row_has_any_value(
            worksheet=worksheet,
            row_index=row_index,
            column_indexes=relevant_columns,
        ):
            last_data_row_index = row_index

    return last_data_row_index


def _row_has_any_value(worksheet, row_index: int, column_indexes: list[int]) -> bool:
    for column_index in column_indexes:
        value = worksheet.cell(row=row_index, column=column_index).value
        if value is None:
            continue
        if str(value).strip():
            return True
    return False


def _select_style_source_row(
    template_sheet: TemplateSheet,
    last_data_row_index: int,
) -> int | None:
    if last_data_row_index >= template_sheet.data_start_row_index:
        return last_data_row_index

    if template_sheet.data_start_row_index > template_sheet.header_row_index:
        return template_sheet.data_start_row_index

    return None


def _copy_row_presentation(
    worksheet,
    source_row_index: int,
    target_row_index: int,
) -> None:
    max_column = worksheet.max_column
    source_row_dimension = worksheet.row_dimensions[source_row_index]
    target_row_dimension = worksheet.row_dimensions[target_row_index]

    if target_row_dimension.height is None and source_row_dimension.height is not None:
        target_row_dimension.height = source_row_dimension.height

    for column_index in range(1, max_column + 1):
        source_cell = worksheet.cell(row=source_row_index, column=column_index)
        target_cell = worksheet.cell(row=target_row_index, column=column_index)

        if source_cell.has_style:
            target_cell._style = copy(source_cell._style)

        if _is_formula(source_cell.value):
            target_cell.value = _translate_formula(
                formula=str(source_cell.value),
                origin=source_cell.coordinate,
                target=target_cell.coordinate,
            )


def _is_formula(value: object) -> bool:
    return isinstance(value, str) and value.startswith("=")


def _translate_formula(formula: str, origin: str, target: str) -> str:
    try:
        return Translator(formula, origin=origin).translate_formula(target)
    except Exception:
        return formula


def _resolve_output_value(
    projected_value: ProjectedTemplateValue,
    worksheet,
    template_sheet: TemplateSheet,
    append_row_index: int,
    last_data_row_index: int,
):
    if projected_value.semantic_key != "row_number":
        return projected_value.value

    previous_value = None
    if last_data_row_index >= template_sheet.data_start_row_index:
        previous_value = worksheet.cell(
            row=last_data_row_index,
            column=projected_value.column_index,
        ).value

    previous_number = _coerce_int(previous_value)
    if previous_number is not None:
        return previous_number + 1

    return append_row_index - template_sheet.data_start_row_index + 1


def _coerce_int(value: object) -> int | None:
    if isinstance(value, bool):
        return None

    if isinstance(value, int):
        return value

    if isinstance(value, float) and value.is_integer():
        return int(value)

    if isinstance(value, str):
        text = value.strip()
        if text.isdigit():
            return int(text)

    return None
