"""레퍼런스 Excel을 `TemplateProfile`로 해석하는 reader 골격."""

from pathlib import Path

from .schema import TemplateColumn, TemplateProfile, TemplateSheet


class TemplateWorkbookReader:
    """기능: 레퍼런스 Excel 워크북을 템플릿 프로필로 해석한다.

    입력:
    - workbook_path: 레퍼런스 Excel 경로
    - profile_id: 사용자 프로필 식별자
    - template_id: 템플릿 식별자
    - max_scan_rows: 헤더 후보를 찾기 위해 앞에서 스캔할 최대 행 수

    반환:
    - reader 인스턴스
    """

    def __init__(
        self,
        workbook_path: str,
        profile_id: str,
        template_id: str = "",
        max_scan_rows: int = 20,
    ) -> None:
        self.workbook_path = str(Path(workbook_path))
        self.profile_id = profile_id
        self.template_id = template_id or Path(workbook_path).stem
        self.max_scan_rows = max_scan_rows

    def read_profile(self) -> TemplateProfile:
        """기능: 워크북에서 `TemplateProfile` 초안을 만든다.

        입력:
        - 없음

        반환:
        - `TemplateProfile` 인스턴스
        """

        workbook = self._load_workbook()
        sheets: list[TemplateSheet] = []

        for worksheet in workbook.worksheets:
            sheet = self._read_sheet(worksheet)
            if sheet is not None:
                sheets.append(sheet)

        notes: list[str] = []
        if not sheets:
            notes.append("visible sheet에서 header row 후보를 찾지 못했다.")

        primary_sheet_name = sheets[0].sheet_name if sheets else None
        return TemplateProfile(
            profile_id=self.profile_id,
            source_workbook_path=self.workbook_path,
            template_id=self.template_id,
            sheets=sheets,
            primary_sheet_name=primary_sheet_name,
            notes=notes,
        )

    def _load_workbook(self):
        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            raise RuntimeError(
                "TemplateWorkbookReader를 사용하려면 `openpyxl`이 필요합니다."
            ) from exc

        return load_workbook(self.workbook_path, data_only=False)

    def _read_sheet(self, worksheet) -> TemplateSheet | None:
        if getattr(worksheet, "sheet_state", "visible") != "visible":
            return None

        header_row_index = self._find_header_row(worksheet)
        if header_row_index is None:
            return None

        columns = self._read_columns(worksheet, header_row_index)
        if not columns:
            return None

        frozen_panes = worksheet.freeze_panes
        if frozen_panes is not None and hasattr(frozen_panes, "coordinate"):
            frozen_panes = frozen_panes.coordinate
        elif frozen_panes is not None:
            frozen_panes = str(frozen_panes)

        return TemplateSheet(
            sheet_name=worksheet.title,
            header_row_index=header_row_index,
            data_start_row_index=header_row_index + 1,
            columns=columns,
            frozen_panes=frozen_panes,
        )

    def _find_header_row(self, worksheet) -> int | None:
        max_row = min(self.max_scan_rows, worksheet.max_row)
        best_row_index: int | None = None
        best_score = 0

        for row_index in range(1, max_row + 1):
            texts = []
            for cell in worksheet[row_index]:
                text = self._stringify(cell.value)
                if text:
                    texts.append(text)

            score = self._score_header_row(texts)
            if score > best_score:
                best_row_index = row_index
                best_score = score

        return best_row_index

    def _score_header_row(self, texts: list[str]) -> int:
        if len(texts) < 2:
            return 0

        unique_count = len(set(texts))
        short_text_count = sum(1 for text in texts if len(text) <= 30)
        return unique_count + short_text_count

    def _read_columns(self, worksheet, header_row_index: int) -> list[TemplateColumn]:
        from openpyxl.utils import get_column_letter

        columns: list[TemplateColumn] = []
        for cell in worksheet[header_row_index]:
            header_text = self._stringify(cell.value)
            if not header_text:
                continue

            example_value, example_cell_ref = self._find_example_value(
                worksheet,
                column_index=cell.column,
                start_row_index=header_row_index + 1,
            )
            columns.append(
                TemplateColumn(
                    header_text=header_text,
                    column_index=cell.column,
                    column_letter=get_column_letter(cell.column),
                    header_cell_ref=cell.coordinate,
                    example_value=example_value,
                    example_cell_ref=example_cell_ref,
                )
            )

        return columns

    def _find_example_value(
        self,
        worksheet,
        column_index: int,
        start_row_index: int,
        max_scan_rows: int = 5,
    ) -> tuple[str | None, str | None]:
        end_row_index = min(worksheet.max_row, start_row_index + max_scan_rows - 1)
        for row_index in range(start_row_index, end_row_index + 1):
            cell = worksheet.cell(row=row_index, column=column_index)
            text = self._stringify(cell.value)
            if text:
                return text, cell.coordinate

        return None, None

    def _stringify(self, value) -> str:
        if value is None:
            return ""
        return str(value).strip()


def read_template_profile(
    workbook_path: str,
    profile_id: str,
    template_id: str = "",
) -> TemplateProfile:
    """기능: 레퍼런스 Excel을 읽어 `TemplateProfile` 초안을 만든다.

    입력:
    - workbook_path: 레퍼런스 Excel 경로
    - profile_id: 사용자 프로필 식별자
    - template_id: 템플릿 식별자

    반환:
    - `TemplateProfile` 인스턴스
    """

    reader = TemplateWorkbookReader(
        workbook_path=workbook_path,
        profile_id=profile_id,
        template_id=template_id,
    )
    return reader.read_profile()
