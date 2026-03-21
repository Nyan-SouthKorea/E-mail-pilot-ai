"""분석 입력용 첨부 자산 요약 helper."""

from __future__ import annotations

import zipfile
from dataclasses import asdict, dataclass
from io import BytesIO
from pathlib import Path

from openpyxl import load_workbook


@dataclass(slots=True)
class ArtifactSummary:
    """기능: 분석 입력에 포함할 첨부 자산 1개 요약을 표현한다.

    입력:
    - evidence_id: LLM 입력에 쓸 근거 id
    - artifact_name: 표시 파일명
    - artifact_kind: 파일 종류
    - summary_text: LLM에 넣을 요약 텍스트

    반환:
    - dataclass 인스턴스
    """

    evidence_id: str
    artifact_name: str
    artifact_kind: str
    summary_text: str

    def to_dict(self) -> dict[str, object]:
        """기능: 요약 객체를 dict로 바꾼다.

        입력:
        - 없음

        반환:
        - 직렬화용 dict
        """

        return asdict(self)


def summarize_attachment_directory(
    attachment_dir: str | Path,
    *,
    artifact_ids: list[str] | None = None,
    start_index: int = 1,
) -> list[ArtifactSummary]:
    """기능: 첨부 디렉토리를 분석 입력용 자산 요약 목록으로 바꾼다.

    입력:
    - attachment_dir: 첨부 디렉토리 경로
    - artifact_ids: 이미 정해진 artifact id 목록. 없으면 순번 기반 생성
    - start_index: 순번 기반 생성 시작값

    반환:
    - `ArtifactSummary` 목록
    """

    root = Path(attachment_dir)
    if not root.exists():
        return []

    file_paths = [path for path in sorted(root.iterdir()) if path.is_file()]
    return summarize_attachment_paths(
        file_paths,
        artifact_ids=artifact_ids,
        start_index=start_index,
    )


def summarize_attachment_paths(
    file_paths: list[Path],
    *,
    artifact_ids: list[str] | None = None,
    start_index: int = 1,
) -> list[ArtifactSummary]:
    """기능: 첨부 파일 경로 목록을 분석 입력용 자산 요약 목록으로 바꾼다.

    입력:
    - file_paths: 첨부 파일 경로 목록
    - artifact_ids: 이미 정해진 artifact id 목록. 없으면 순번 기반 생성
    - start_index: 순번 기반 생성 시작값

    반환:
    - `ArtifactSummary` 목록
    """

    summaries: list[ArtifactSummary] = []
    for offset, path in enumerate(file_paths):
        evidence_id = _resolve_artifact_id(
            artifact_ids=artifact_ids,
            offset=offset,
            start_index=start_index,
        )
        if path.suffix.lower() == ".zip":
            summaries.extend(_summarize_zip_artifacts(path, evidence_id=evidence_id))
            continue

        summaries.append(
            ArtifactSummary(
                evidence_id=evidence_id,
                artifact_name=path.name,
                artifact_kind=path.suffix.lower().lstrip(".") or "file",
                summary_text=f"첨부파일: {path.name}",
            )
        )

    return summaries


def _resolve_artifact_id(
    *,
    artifact_ids: list[str] | None,
    offset: int,
    start_index: int,
) -> str:
    if artifact_ids and offset < len(artifact_ids):
        return artifact_ids[offset]
    return f"artifact_{start_index + offset}"


def _summarize_zip_artifacts(path: Path, *, evidence_id: str) -> list[ArtifactSummary]:
    summaries: list[ArtifactSummary] = []
    with zipfile.ZipFile(path) as archive:
        names = archive.namelist()
        for index, name in enumerate(names, start=1):
            lower_name = name.lower()
            nested_evidence_id = evidence_id if len(names) == 1 else f"{evidence_id}_{index}"
            if lower_name.endswith(".xlsx"):
                summary_text = _summarize_xlsx_from_zip(archive, name)
                artifact_kind = "xlsx"
            elif lower_name.endswith(".pdf"):
                summary_text = f"ZIP 내부 PDF 파일: {name}"
                artifact_kind = "pdf"
            else:
                summary_text = f"ZIP 내부 파일: {name}"
                artifact_kind = Path(name).suffix.lower().lstrip(".") or "file"

            summaries.append(
                ArtifactSummary(
                    evidence_id=nested_evidence_id,
                    artifact_name=name,
                    artifact_kind=artifact_kind,
                    summary_text=summary_text,
                )
            )
    return summaries


def _summarize_xlsx_from_zip(archive: zipfile.ZipFile, name: str) -> str:
    workbook = load_workbook(filename=BytesIO(archive.read(name)), data_only=True)
    worksheet = workbook[workbook.sheetnames[0]]

    lines = [f"ZIP 내부 XLSX 파일: {name}"]
    captured_rows = 0
    for row in worksheet.iter_rows(min_row=1, max_row=24, values_only=True):
        values = [str(value).strip() for value in row if value not in (None, "")]
        if not values:
            continue
        lines.append(" | ".join(values))
        captured_rows += 1
        if captured_rows >= 12:
            break

    return "\n".join(lines)
