"""전체 runtime bundle을 triage하고 review board를 생성하는 smoke."""

from __future__ import annotations

import argparse
import json
import os
import sys
from dataclasses import asdict, dataclass, field
from datetime import datetime
from html import escape
from pathlib import Path
from urllib.parse import quote

if __package__ in (None, ""):
    sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

from llm import OpenAIResponsesConfig, OpenAIResponsesWrapper
from mailbox.bundle_reader import (
    list_valid_runtime_bundle_directories,
    load_normalized_message_from_bundle,
)
from project_paths import ProfilePaths, default_example_profile_root

if __package__ in (None, ""):
    from analysis.materialized_bundle_pipeline_smoke import default_template_path
    from analysis.materialized_bundle_smoke import run_materialized_bundle_analysis_smoke
    from analysis.schema import ExtractedField, ExtractedRecord
else:
    from .materialized_bundle_pipeline_smoke import default_template_path
    from .materialized_bundle_smoke import run_materialized_bundle_analysis_smoke
    from .schema import ExtractedField, ExtractedRecord

from exports import (
    append_projected_row_to_workbook,
    apply_hybrid_template_mapping,
    build_timestamped_export_workbook_path,
    project_record_to_template,
    read_template_profile,
)
from openpyxl import load_workbook


@dataclass(slots=True)
class InboxReviewBoardItem:
    """기능: review board의 메일 1건 행 데이터를 표현한다."""

    bundle_id: str
    bundle_root: str
    received_at: str | None
    sender: str
    subject: str
    attachment_count: int
    triage_label: str
    triage_reason: str
    triage_confidence: float | None
    analysis_source: str
    export_status: str
    appended_row_index: int | None = None
    unresolved_columns: list[str] = field(default_factory=list)
    company_name: str = ""
    contact_name: str = ""
    email_address: str = ""
    application_purpose: str = ""
    request_summary: str = ""
    extracted_record_path: str | None = None
    projected_row_path: str | None = None
    summary_path: str | None = None
    preview_path: str | None = None
    notes: list[str] = field(default_factory=list)

    def to_dict(self) -> dict[str, object]:
        return asdict(self)


@dataclass(slots=True)
class InboxReviewBoardReport:
    """기능: 전체 batch review와 HTML board 생성 결과를 표현한다."""

    generated_at: str
    profile_root: str
    review_json_path: str
    review_html_path: str
    output_workbook_path: str
    total_bundle_count: int
    application_count: int
    not_application_count: int
    needs_human_review_count: int
    exported_count: int
    skipped_count: int
    failed_count: int
    items: list[InboxReviewBoardItem] = field(default_factory=list)
    notes: list[str] = field(default_factory=list)

    def to_dict(self) -> dict[str, object]:
        payload = asdict(self)
        payload["items"] = [item.to_dict() for item in self.items]
        return payload


@dataclass(slots=True)
class HeuristicTriageDecision:
    """기능: bundle prefilter triage 결정을 표현한다."""

    should_run_llm: bool
    triage_label: str
    triage_reason: str
    triage_confidence: float


def default_profile_root() -> Path:
    """기능: 현재 예시 사용자 프로필 루트 기본 경로를 반환한다."""

    return default_example_profile_root()


def run_inbox_review_board_smoke(
    *,
    profile_id: str,
    profile_root: str,
    template_path: str,
    bundle_root: str | None = None,
    bundle_limit: int | None = None,
    reuse_existing_analysis: bool = False,
    wrapper: OpenAIResponsesWrapper | None = None,
) -> InboxReviewBoardReport:
    """기능: runtime bundle 전체를 triage하고 review board를 생성한다."""

    profile_paths = ProfilePaths(profile_root)
    profile_paths.ensure_runtime_dirs()

    wrapper = wrapper or OpenAIResponsesWrapper(
        OpenAIResponsesConfig(
            usage_log_path=str(profile_paths.llm_usage_log_path()),
        )
    )

    template_profile = read_template_profile(
        workbook_path=template_path,
        profile_id=profile_id,
        template_id=Path(template_path).stem,
    )
    mapped_profile, mapping = apply_hybrid_template_mapping(
        template_profile,
        wrapper=wrapper,
    )

    review_root = profile_paths.runtime_review_logs_root()
    review_root.mkdir(parents=True, exist_ok=True)
    timestamp_label = datetime.now().strftime("%y%m%d_%H%M")
    review_json_path = review_root / f"{timestamp_label}_inbox_review_board.json"
    review_html_path = review_root / f"{timestamp_label}_inbox_review_board.html"
    workbook_output = build_timestamped_export_workbook_path(
        profile_root=profile_root,
        template_workbook_path=template_path,
    )
    if workbook_output.exists():
        workbook_output.unlink()
    _prepare_empty_output_workbook(
        template_path=Path(template_path),
        output_workbook_path=workbook_output,
        template_profile=mapped_profile,
    )

    notes: list[str] = []
    if bundle_root:
        bundle_directories = [Path(bundle_root)]
    else:
        bundle_directories = list_valid_runtime_bundle_directories(profile_root)
        if bundle_limit is not None and bundle_limit > 0:
            bundle_directories = sorted(
                bundle_directories,
                key=lambda item: item.name,
                reverse=True,
            )[:bundle_limit]
            notes.append(f"review board는 최근 {bundle_limit}개 bundle만 빠른 테스트 범위로 재생성했다.")

    items: list[InboxReviewBoardItem] = []
    exported_count = 0
    failed_count = 0
    notes.append("전량 review board에는 모든 valid bundle을 포함하고, 첨부/신청 신호가 없는 메일은 heuristic prefilter로 비신청 선분류했다.")

    if mapping.unresolved_headers:
        notes.append("일부 템플릿 헤더는 rule 기반으로 확정되지 않아 LLM fallback 결과를 사용했다.")

    llm_full_analysis_count = 0
    heuristic_prefilter_count = 0

    for index, directory in enumerate(bundle_directories, start=1):
        normalized = load_normalized_message_from_bundle(directory)
        try:
            pretriage = _build_prefilter_triage_decision(normalized)
            if pretriage.should_run_llm:
                llm_full_analysis_count += 1
                analysis_results = run_materialized_bundle_analysis_smoke(
                    profile_root=profile_root,
                    bundle_root=str(directory),
                    reuse_existing_analysis=reuse_existing_analysis,
                    wrapper=wrapper,
                )
                if not analysis_results:
                    raise RuntimeError("bundle 분석 결과를 만들지 못했다.")
                analysis_result = analysis_results[0]
                extracted_record = ExtractedRecord.from_dict(
                    json.loads(
                        Path(analysis_result.extracted_record_path).read_text(encoding="utf-8")
                    )
                )
                analysis_source = "llm_full_analysis"
                item_notes = list(analysis_result.notes)
            else:
                extracted_record = _build_prefilter_triage_record(
                    normalized_message=normalized,
                    decision=pretriage,
                )
                extracted_record_path = (
                    profile_paths.runtime_analysis_logs_root()
                    / f"{normalized.bundle_id}_extracted_record.json"
                )
                extracted_record_path.parent.mkdir(parents=True, exist_ok=True)
                extracted_record_path.write_text(
                    json.dumps(extracted_record.to_dict(), ensure_ascii=False, indent=2),
                    encoding="utf-8",
                )
                analysis_result = type(
                    "HeuristicAnalysisResult",
                    (),
                    {
                        "extracted_record_path": str(extracted_record_path),
                        "notes": [pretriage.triage_reason],
                    },
                )()
                analysis_source = "heuristic_prefilter"
                heuristic_prefilter_count += 1
                item_notes = list(analysis_result.notes)

            projected_row = project_record_to_template(mapped_profile, extracted_record)
            projected_row_path = (
                profile_paths.runtime_exports_logs_root()
                / f"{normalized.bundle_id}_projected_row.json"
            )
            projected_row_path.parent.mkdir(parents=True, exist_ok=True)
            projected_row_path.write_text(
                json.dumps(projected_row.to_dict(), ensure_ascii=False, indent=2),
                encoding="utf-8",
            )

            export_status = _resolve_export_status(extracted_record)
            appended_row_index: int | None = None
            item_notes.extend(projected_row.notes)
            if export_status == "exported":
                append_result = append_projected_row_to_workbook(
                    profile=mapped_profile,
                    projected_row=projected_row,
                    output_workbook_path=str(workbook_output),
                    source_workbook_path=str(template_path),
                )
                exported_count += 1
                appended_row_index = append_result.appended_row_index
                item_notes.extend(append_result.notes)

            field_map = extracted_record.field_map()
            items.append(
                InboxReviewBoardItem(
                    bundle_id=normalized.bundle_id,
                    bundle_root=str(directory),
                    received_at=normalized.received_at,
                    sender=_format_address(normalized.sender),
                    subject=normalized.subject,
                    attachment_count=len(normalized.attachment_artifact_ids),
                    triage_label=extracted_record.triage_label,
                    triage_reason=extracted_record.triage_reason,
                    triage_confidence=extracted_record.triage_confidence,
                    analysis_source=analysis_source,
                    export_status=export_status,
                    appended_row_index=appended_row_index,
                    unresolved_columns=list(projected_row.unresolved_columns),
                    company_name=_field_value(field_map, "company_name"),
                    contact_name=_field_value(field_map, "contact_name"),
                    email_address=_field_value(field_map, "email_address"),
                    application_purpose=_field_value(field_map, "application_purpose"),
                    request_summary=_field_value(field_map, "request_summary"),
                    extracted_record_path=analysis_result.extracted_record_path,
                    projected_row_path=str(projected_row_path),
                    summary_path=str(directory / "summary.md"),
                    preview_path=str(directory / "preview.html"),
                    notes=item_notes,
                )
            )
        except Exception as exc:
            failed_count += 1
            items.append(
                InboxReviewBoardItem(
                    bundle_id=normalized.bundle_id,
                    bundle_root=str(directory),
                    received_at=normalized.received_at,
                    sender=_format_address(normalized.sender),
                    subject=normalized.subject,
                    attachment_count=len(normalized.attachment_artifact_ids),
                    triage_label="needs_human_review",
                    triage_reason="분석 또는 projection 단계가 실패해 수동 확인이 필요하다.",
                    triage_confidence=0.0,
                    analysis_source="failed_before_analysis",
                    export_status="failed",
                    summary_path=str(directory / "summary.md"),
                    preview_path=str(directory / "preview.html"),
                    notes=[f"{exc.__class__.__name__}: {exc}"],
                )
            )
        if index % 100 == 0 or index == len(bundle_directories):
            print(
                (
                    f"[inbox_review] {index}/{len(bundle_directories)} processed | "
                    f"exported={exported_count} failed={failed_count} "
                    f"llm={llm_full_analysis_count} heuristic={heuristic_prefilter_count}"
                ),
                file=sys.stderr,
                flush=True,
            )

    if not exported_count:
        notes.append("자동 export 대상 신청서가 없어 reference template을 그대로 복사한 빈 결과 workbook을 남겼다.")
    notes.append(
        (
            f"analysis_source_counts: llm_full_analysis={llm_full_analysis_count}, "
            f"heuristic_prefilter={heuristic_prefilter_count}"
        )
    )

    items.sort(
        key=lambda item: (
            item.received_at or "",
            item.bundle_id,
        ),
        reverse=True,
    )

    application_count = sum(1 for item in items if item.triage_label == "application")
    not_application_count = sum(
        1 for item in items if item.triage_label == "not_application"
    )
    needs_human_review_count = sum(
        1 for item in items if item.triage_label == "needs_human_review"
    )
    skipped_count = sum(
        1
        for item in items
        if item.export_status.startswith("skipped_")
    )

    report = InboxReviewBoardReport(
        generated_at=datetime.now().isoformat(),
        profile_root=profile_root,
        review_json_path=str(review_json_path),
        review_html_path=str(review_html_path),
        output_workbook_path=str(workbook_output),
        total_bundle_count=len(items),
        application_count=application_count,
        not_application_count=not_application_count,
        needs_human_review_count=needs_human_review_count,
        exported_count=exported_count,
        skipped_count=skipped_count,
        failed_count=failed_count,
        items=items,
        notes=notes,
    )
    review_json_path.write_text(
        json.dumps(report.to_dict(), ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    review_html_path.write_text(
        build_inbox_review_board_html(report=report),
        encoding="utf-8",
    )
    return report


def build_inbox_review_board_html(*, report: InboxReviewBoardReport) -> str:
    """기능: 정적 HTML review board를 생성한다."""

    review_root = Path(report.review_html_path).parent
    rows: list[str] = []
    for item in report.items:
        triage_class = f"triage-{item.triage_label}"
        export_class = "exported" if item.export_status == "exported" else "skipped"
        rows.append(
            "<tr>"
            f"<td>{escape(item.received_at or '')}</td>"
            f"<td>{escape(item.sender)}</td>"
            f"<td>{escape(item.subject or '(제목 없음)')}</td>"
            f"<td>{item.attachment_count}</td>"
            f"<td><span class=\"badge {triage_class}\">{escape(item.triage_label)}</span></td>"
            f"<td>{escape(item.analysis_source)}</td>"
            f"<td>{escape(item.triage_reason)}</td>"
            f"<td>{'' if item.triage_confidence is None else f'{item.triage_confidence:.2f}'}</td>"
            f"<td><span class=\"badge export-{export_class}\">{escape(item.export_status)}</span></td>"
            f"<td>{len(item.unresolved_columns)}</td>"
            f"<td>{escape(_join_non_empty([item.company_name, item.contact_name, item.email_address]))}</td>"
            f"<td>{escape(item.application_purpose)}</td>"
            f"<td>{escape(item.request_summary)}</td>"
            f"<td>{_build_link_list_html(review_root, item)}</td>"
            "</tr>"
        )

    notes_html = "".join(f"<li>{escape(note)}</li>" for note in report.notes) or "<li>(없음)</li>"
    return (
        "<!doctype html>\n"
        "<html lang=\"ko\">\n"
        "<head>\n"
        "  <meta charset=\"utf-8\">\n"
        "  <title>Inbox Review Board</title>\n"
        "  <style>\n"
        "    body{font-family:'Malgun Gothic',sans-serif;background:#f6f7f9;color:#1c1f24;margin:24px;}\n"
        "    h1{margin:0 0 8px;}\n"
        "    .sub{margin:0 0 20px;color:#556070;}\n"
        "    .cards{display:grid;grid-template-columns:repeat(6,minmax(120px,1fr));gap:12px;margin:20px 0;}\n"
        "    .card{background:#fff;border:1px solid #dbe1e8;border-radius:12px;padding:14px 16px;box-shadow:0 1px 3px rgba(0,0,0,.04);}\n"
        "    .card .label{display:block;font-size:12px;color:#667085;margin-bottom:6px;}\n"
        "    .card .value{font-size:24px;font-weight:700;}\n"
        "    table{width:100%;border-collapse:collapse;background:#fff;border:1px solid #dbe1e8;}\n"
        "    th,td{padding:10px 12px;border-bottom:1px solid #edf1f5;vertical-align:top;text-align:left;font-size:13px;}\n"
        "    th{background:#f8fafc;position:sticky;top:0;z-index:1;}\n"
        "    .badge{display:inline-block;padding:4px 8px;border-radius:999px;font-size:12px;font-weight:700;}\n"
        "    .triage-application{background:#dcfce7;color:#166534;}\n"
        "    .triage-not_application{background:#e5e7eb;color:#374151;}\n"
        "    .triage-needs_human_review{background:#fef3c7;color:#92400e;}\n"
        "    .export-exported{background:#dbeafe;color:#1d4ed8;}\n"
        "    .export-skipped{background:#f3f4f6;color:#4b5563;}\n"
        "    .links a{display:inline-block;margin-right:8px;margin-bottom:4px;}\n"
        "    .notes{margin-top:20px;background:#fff;border:1px solid #dbe1e8;border-radius:12px;padding:16px;}\n"
        "  </style>\n"
        "</head>\n"
        "<body>\n"
        "  <h1>Inbox Review Board</h1>\n"
        f"  <p class=\"sub\">generated_at: {escape(report.generated_at)} | workbook: {escape(Path(report.output_workbook_path).name)}</p>\n"
        "  <section class=\"cards\">\n"
        f"    <div class=\"card\"><span class=\"label\">총 메일 수</span><span class=\"value\">{report.total_bundle_count}</span></div>\n"
        f"    <div class=\"card\"><span class=\"label\">application</span><span class=\"value\">{report.application_count}</span></div>\n"
        f"    <div class=\"card\"><span class=\"label\">not_application</span><span class=\"value\">{report.not_application_count}</span></div>\n"
        f"    <div class=\"card\"><span class=\"label\">needs_human_review</span><span class=\"value\">{report.needs_human_review_count}</span></div>\n"
        f"    <div class=\"card\"><span class=\"label\">exported</span><span class=\"value\">{report.exported_count}</span></div>\n"
        f"    <div class=\"card\"><span class=\"label\">failed</span><span class=\"value\">{report.failed_count}</span></div>\n"
        "  </section>\n"
        "  <table>\n"
        "    <thead>\n"
        "      <tr>\n"
        "        <th>received_at</th><th>sender</th><th>subject</th><th>attach</th><th>triage</th><th>source</th><th>reason</th><th>conf</th><th>export</th><th>unresolved</th><th>key fields</th><th>application purpose</th><th>request summary</th><th>links</th>\n"
        "      </tr>\n"
        "    </thead>\n"
        "    <tbody>\n"
        f"{''.join(rows)}\n"
        "    </tbody>\n"
        "  </table>\n"
        "  <section class=\"notes\">\n"
        "    <h2>Run Notes</h2>\n"
        f"    <ul>{notes_html}</ul>\n"
        "  </section>\n"
        "</body>\n"
        "</html>\n"
    )


def _build_link_list_html(review_root: Path, item: InboxReviewBoardItem) -> str:
    links: list[str] = []
    for label, target in [
        ("summary", item.summary_path),
        ("preview", item.preview_path),
        ("record", item.extracted_record_path),
        ("projected", item.projected_row_path),
    ]:
        if not target:
            continue
        href = _relative_href(review_root, Path(target))
        links.append(f"<a href=\"{href}\">{escape(label)}</a>")
    return f"<div class=\"links\">{''.join(links) or '(없음)'}</div>"


def _relative_href(base_dir: Path, target_path: Path) -> str:
    relative = Path(os.path.relpath(target_path, base_dir))
    return "/".join(quote(part) for part in relative.parts)


def _field_value(field_map: dict[str, object], field_name: str) -> str:
    field = field_map.get(field_name)
    if field is None:
        return ""
    normalized_value = getattr(field, "normalized_value", None)
    value = getattr(field, "value", "")
    return (normalized_value or value or "").strip()


def _format_address(address) -> str:
    if getattr(address, "name", None):
        return f"{address.name} <{address.email}>"
    return address.email


def _join_non_empty(values: list[str]) -> str:
    return " | ".join(value for value in values if value.strip())


def _prepare_empty_output_workbook(
    *,
    template_path: Path,
    output_workbook_path: Path,
    template_profile,
) -> None:
    workbook = load_workbook(template_path)
    for sheet in template_profile.sheets:
        worksheet = workbook[sheet.sheet_name]
        for row_index in range(sheet.data_start_row_index, worksheet.max_row + 1):
            for column_index in range(1, worksheet.max_column + 1):
                cell = worksheet.cell(row=row_index, column=column_index)
                cell.value = None
                cell.hyperlink = None

    output_workbook_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_workbook_path)


def _build_prefilter_triage_decision(normalized_message) -> HeuristicTriageDecision:
    subject_text = (normalized_message.subject or "").casefold()
    body_text = (normalized_message.body_text or "").casefold()
    sender_email = (normalized_message.sender.email or "").casefold()
    recipient_emails = {
        address.email.casefold()
        for address in normalized_message.to + normalized_message.cc + normalized_message.bcc
        if address.email
    }

    subject_tokens = [
        "신청",
        "모집",
        "참가",
        "지원",
        "application",
        "registration",
    ]
    strong_body_tokens = [
        "참가 신청",
        "신청 드립니다",
        "신청합니다",
        "지원합니다",
        "지원서 제출",
        "submit application",
    ]
    has_attachment = bool(normalized_message.attachment_artifact_ids)
    has_subject_signal = any(token in subject_text for token in subject_tokens)
    has_strong_body_signal = any(token in body_text for token in strong_body_tokens)
    is_self_mail = sender_email and sender_email in recipient_emails
    is_internal_sender = "@kova.or.kr" in sender_email
    is_system_sender = any(
        token in sender_email
        for token in ["newsletter", "antispam", "wblock", "@system"]
    )

    if is_system_sender:
        return HeuristicTriageDecision(
            should_run_llm=False,
            triage_label="not_application",
            triage_reason="시스템 알림 또는 뉴스레터 발신자로 보여 비신청으로 선분류했다.",
            triage_confidence=0.97,
        )

    if (
        not is_internal_sender
        and ((has_attachment and (has_subject_signal or has_strong_body_signal)) or has_strong_body_signal)
    ):
        return HeuristicTriageDecision(
            should_run_llm=True,
            triage_label="needs_human_review",
            triage_reason="첨부와 신청 신호가 함께 있거나 본문에 강한 신청 표현이 있어 full analysis 대상으로 유지했다.",
            triage_confidence=0.67,
        )

    if is_self_mail:
        return HeuristicTriageDecision(
            should_run_llm=False,
            triage_label="not_application",
            triage_reason="보낸사람과 받는사람이 겹치는 내부 정리 메일로 보여 비신청으로 선분류했다.",
            triage_confidence=0.94,
        )

    return HeuristicTriageDecision(
        should_run_llm=False,
        triage_label="not_application",
        triage_reason="첨부와 명시적 신청 신호 조합이 없어 batch prefilter에서 비신청으로 선분류했다.",
        triage_confidence=0.78,
    )


def _build_prefilter_triage_record(
    *,
    normalized_message,
    decision: HeuristicTriageDecision,
) -> ExtractedRecord:
    field_map = []
    if normalized_message.sender.name:
        field_map.append(
            {
                "field_name": "contact_name",
                "value": normalized_message.sender.name,
                "normalized_value": normalized_message.sender.name,
            }
        )
    if normalized_message.sender.email:
        field_map.append(
            {
                "field_name": "email_address",
                "value": normalized_message.sender.email,
                "normalized_value": normalized_message.sender.email,
            }
        )

    fields = [
        {
            "field_name": item["field_name"],
            "value": item["value"],
            "normalized_value": item["normalized_value"],
        }
        for item in field_map
    ]

    record = ExtractedRecord(
        bundle_id=normalized_message.bundle_id,
        message_key=normalized_message.message_key,
        record_type="email_message",
        category="prefilter_triage",
        fields=[
            ExtractedField(
                field_name=item["field_name"],
                value=item["value"],
                normalized_value=item["normalized_value"],
                confidence=decision.triage_confidence,
                evidence_ids=[],
                notes="LLM full analysis 전에 prefilter triage용으로 채운 최소 필드다.",
            )
            for item in fields
        ],
        summary_one_line=(normalized_message.subject or "비신청 메일").strip(),
        summary_short=_build_prefilter_summary(normalized_message),
        triage_label=decision.triage_label,
        triage_reason=decision.triage_reason,
        triage_confidence=decision.triage_confidence,
        overall_confidence=decision.triage_confidence,
        action_hints=[],
        unresolved_questions=[],
        source_artifact_ids=list(normalized_message.attachment_artifact_ids),
    )
    return record


def _build_prefilter_summary(normalized_message) -> str:
    body_lines = [
        line.strip()
        for line in (normalized_message.body_text or "").splitlines()
        if line.strip()
    ]
    preview = " ".join(body_lines[:2]).strip()
    if preview:
        return preview[:220]
    return (normalized_message.subject or "본문 미리보기 없음").strip()


def _resolve_export_status(record: ExtractedRecord) -> str:
    if record.triage_label != "application":
        if record.triage_label == "not_application":
            return "skipped_not_application"
        return "skipped_needs_human_review"

    if not _has_export_gate_signals(record):
        return "skipped_missing_export_signals"
    return "exported"


def _has_export_gate_signals(record: ExtractedRecord) -> bool:
    field_map = record.field_map()
    has_company_signal = bool(_field_value(field_map, "company_name"))
    has_contact_signal = any(
        _field_value(field_map, field_name)
        for field_name in ["contact_name", "phone_number", "email_address"]
    )
    return has_company_signal and has_contact_signal


def main() -> None:
    """기능: CLI에서 전체 inbox review board smoke를 실행한다."""

    parser = argparse.ArgumentParser(description="runtime inbox review board smoke")
    parser.add_argument("--profile-id", default="kim-jm")
    parser.add_argument("--profile-root", default=str(default_profile_root()))
    parser.add_argument("--template-path", default=str(default_template_path()))
    parser.add_argument("--bundle-root")
    parser.add_argument(
        "--reuse-existing-analysis",
        action="store_true",
        help="이미 저장된 extracted_record JSON이 있으면 재사용한다.",
    )
    args = parser.parse_args()

    report = run_inbox_review_board_smoke(
        profile_id=args.profile_id,
        profile_root=args.profile_root,
        template_path=args.template_path,
        bundle_root=args.bundle_root,
        reuse_existing_analysis=args.reuse_existing_analysis,
    )
    print(json.dumps(report.to_dict(), ensure_ascii=False, indent=2))
    if report.failed_count:
        raise SystemExit(1)


if __name__ == "__main__":
    main()
